#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
績效考核評分表自動生成腳本
115年度 上半年 績效考核 — 樂聯工業股份有限公司

執行方式：python3 generate_eval_sheets.py
輸出：eval_output/ 資料夾，每人一個 Excel 評分表
"""

import os
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle

# ============================================================
# 路徑設定
# ============================================================
TEMPLATE_FILE = '/home/user/website/115年終考核題目_修改題目對照.xlsm'
OUTPUT_DIR    = '/home/user/website/eval_output'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ============================================================
# 不參與本次評核的人員（新進人員）
# ============================================================
EXCLUDED = {'陳明軒', '蔡坤惠', '蔡政博', '張永盛', '陳孟晴', '潘佩君'}  # 含育嬰假

# ============================================================
# 員工資料表
# role 說明：
#   admin_staff         = 一般行政人員（助理、專員、工程師等）
#   admin_group_leader  = 行政組長（物管組長、採購組長、成倉組長、業務課組長）
#   section_chief       = 課長/副課長
#   deputy_manager      = 副理
#   manager             = 經理
#   field_worker        = 現場作業員
#   field_leader        = 現場班長
#   field_group_leader  = 現場組長/副組長
#   top_mgmt            = 副總/總經理（不參與一般互評）
#
# cross_group = 跨部門共評分組（admin1 或 admin2）
#   admin1 = 廠務/資材/品保/研發/營業 側（工廠相關行政）
#   admin2 = 財務/總經理室/業務/總務 側（公司管理行政）
# ============================================================

EMPLOYEES = {
    # ─── 業務課 ───────────────────────────────────────────────
    '伍珈彣': {'dept': '業務課', 'role': 'admin_group_leader', 'cross_group': 'admin2'},
    '鄭淑鐘': {'dept': '業務課', 'role': 'admin_staff',        'cross_group': 'admin2'},
    '陳玉慧': {'dept': '業務課', 'role': 'admin_staff',        'cross_group': 'admin2'},

    # ─── 研發課 ───────────────────────────────────────────────
    # 白惇維：研發課課長 + 加工組代理組長（雙重職責，特殊處理）
    '白惇維': {'dept': '研發課', 'role': 'section_chief', 'cross_group': 'admin1',
               'extra_field_group': '加工組'},  # 額外負責加工組初評
    '王柳琦': {'dept': '研發課', 'role': 'admin_staff',   'cross_group': 'admin1'},
    '徐惠蘭': {'dept': '研發課', 'role': 'admin_staff',   'cross_group': 'admin1'},
    '沙洋':   {'dept': '研發課', 'role': 'admin_staff',   'cross_group': 'admin1'},

    # ─── 總務課 ───────────────────────────────────────────────
    '張欽發': {'dept': '總務課', 'role': 'admin_staff', 'cross_group': 'admin2'},
    '陳福郎': {'dept': '總務課', 'role': 'admin_staff', 'cross_group': 'admin2'},
    '吳桂梅': {'dept': '總務課', 'role': 'admin_staff', 'cross_group': 'admin2'},
    '蔡欣芸': {'dept': '總務課', 'role': 'admin_staff', 'cross_group': 'admin2'},

    # ─── 財務部 ───────────────────────────────────────────────
    '孫素琴': {'dept': '財務部', 'role': 'admin_staff', 'cross_group': 'admin2'},
    '王美婷': {'dept': '財務部', 'role': 'admin_staff', 'cross_group': 'admin2'},
    '陳燕華': {'dept': '財務部', 'role': 'admin_staff', 'cross_group': 'admin2'},

    # ─── 品保課 ───────────────────────────────────────────────
    '朱冠瑋': {'dept': '品保課', 'role': 'section_chief', 'cross_group': 'admin1'},
    '周清家': {'dept': '品保課', 'role': 'admin_staff',   'cross_group': 'admin1'},
    '詹郁瑛': {'dept': '品保課', 'role': 'admin_staff',   'cross_group': 'admin1'},
    '林依婷': {'dept': '品保課', 'role': 'admin_staff',   'cross_group': 'admin1'},

    # ─── 廠務部（廠務室工程師）────────────────────────────────
    '曾夷璋': {'dept': '廠務部', 'role': 'engineer', 'cross_group': None},
    '陳家祥': {'dept': '廠務部', 'role': 'engineer', 'cross_group': None},

    # ─── 資材課（行政人員）────────────────────────────────────
    '王秀雯': {'dept': '資材課', 'role': 'section_chief',       'cross_group': 'admin1'},
    '石容瑄': {'dept': '資材課', 'role': 'admin_group_leader',  'cross_group': 'admin1'},
    '陳惠萍': {'dept': '資材課', 'role': 'admin_group_leader',  'cross_group': 'admin1'},
    '曾泓霖': {'dept': '成倉組', 'role': 'field_group_leader', 'cross_group': None},  # 成倉組長，歸現場
    '尤念萍': {'dept': '資材課', 'role': 'admin_staff',         'cross_group': 'admin1'},
    '林德和': {'dept': '資材課', 'role': 'admin_staff',         'cross_group': 'admin1'},
    '羅文彥': {'dept': '資材課', 'role': 'admin_staff',         'cross_group': 'admin1'},
    '楊心恬': {'dept': '資材課', 'role': 'admin_staff',         'cross_group': 'admin1'},
    '陳巧倫': {'dept': '資材課', 'role': 'admin_staff',         'cross_group': 'admin1'},
    '郭惠婷': {'dept': '資材課', 'role': 'admin_staff',         'cross_group': 'admin1'},
    '尤秐棠': {'dept': '資材課', 'role': 'admin_staff',         'cross_group': 'admin1'},
    '陳孟怡': {'dept': '資材課', 'role': 'admin_staff',         'cross_group': 'admin1'},
    '戴曜瑋': {'dept': '成倉組', 'role': 'field_worker',        'cross_group': None},
    '清馬拉': {'dept': '成倉組', 'role': 'field_worker',        'cross_group': None},
    '沙達':   {'dept': '成倉組', 'role': 'field_worker',        'cross_group': None},

    # ─── 營業部 ───────────────────────────────────────────────
    '劉祐誠': {'dept': '營業部', 'role': 'manager', 'cross_group': 'admin1'},

    # ─── 總經理室 ─────────────────────────────────────────────
    '龔淑屏': {'dept': '總經理室', 'role': 'deputy_manager', 'cross_group': 'admin2'},
    '鄭景文': {'dept': '總經理室', 'role': 'top_mgmt',       'cross_group': 'admin2'},
    '陳佳倫': {'dept': '總經理室', 'role': 'top_mgmt',       'cross_group': 'admin2'},
    '黃士咸': {'dept': '總經理室', 'role': 'admin_staff',    'cross_group': 'admin2'},
    '尹詩婷': {'dept': '總經理室', 'role': 'admin_staff',    'cross_group': 'admin2'},
    '李庭慧': {'dept': '總經理室', 'role': 'admin_staff',    'cross_group': 'admin2'},

    # ─── 現場：塗裝組 ─────────────────────────────────────────
    '郭惠燕': {'dept': '塗裝組', 'role': 'field_group_leader', 'cross_group': None},
    '周安琪': {'dept': '塗裝組', 'role': 'field_leader',       'cross_group': None},
    '威耐':   {'dept': '塗裝組', 'role': 'field_leader',       'cross_group': None},
    '賴慶燁': {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '夏塔拉': {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '那孔':   {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '肅坦':   {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '百翔':   {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '那彭':   {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '阿迪朋': {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '阿沙文': {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '宇藍南': {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '松沙':   {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '南龍裡': {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '南天鵬': {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '陳俊智': {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '磊瓦':   {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},
    '黃氏金莉': {'dept': '塗裝組', 'role': 'field_worker',     'cross_group': None},
    '陳世傑': {'dept': '塗裝組', 'role': 'field_worker',       'cross_group': None},

    # ─── 現場：組一組 ─────────────────────────────────────────
    '陳怡珊': {'dept': '組一組', 'role': 'field_group_leader', 'cross_group': None},
    '馬程鳳': {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '連維敏': {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '陳意琛': {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '松甲':   {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '朱淑君': {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '蘇拉岡': {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '周柏杰': {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '呂雅秀': {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '朗吾':   {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '巴林亞': {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '及可':   {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},
    '吉沙達': {'dept': '組一組', 'role': 'field_worker',       'cross_group': None},

    # ─── 現場：組二組 ─────────────────────────────────────────
    '吉實達': {'dept': '組二組', 'role': 'field_group_leader', 'cross_group': None},
    '紀佩欣': {'dept': '組二組', 'role': 'field_group_leader', 'cross_group': None},
    '潘盈蓁': {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},
    '姜威藍': {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},
    '吳拉查': {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},
    '薩財':   {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},
    '隆峽':   {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},
    '三理':   {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},
    '郭藍林': {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},
    '陳凱':   {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},
    '查朋':   {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},
    '提拉瓦': {'dept': '組二組', 'role': 'field_worker',       'cross_group': None},

    # ─── 現場：組三組（蔡坤惠排除）───────────────────────────
    '蔡麗娟': {'dept': '組三組', 'role': 'field_worker', 'cross_group': None},
    '梁美麗': {'dept': '組三組', 'role': 'field_worker', 'cross_group': None},
    '林綉玲': {'dept': '組三組', 'role': 'field_worker', 'cross_group': None},
    '郭韋霆': {'dept': '組三組', 'role': 'field_worker', 'cross_group': None},
    '莊慈君': {'dept': '組三組', 'role': 'field_worker', 'cross_group': None},
    '阿匹雷': {'dept': '組三組', 'role': 'field_worker', 'cross_group': None},
    '拉匹潘': {'dept': '組三組', 'role': 'field_worker', 'cross_group': None},
    '阮美和': {'dept': '組三組', 'role': 'field_worker', 'cross_group': None},
    '拉桑':   {'dept': '組三組', 'role': 'field_worker', 'cross_group': None},

    # ─── 現場：沖床組 ─────────────────────────────────────────
    '覃麗嬋': {'dept': '沖床組', 'role': 'field_group_leader', 'cross_group': None},
    '李昱鋐': {'dept': '沖床組', 'role': 'field_leader',       'cross_group': None},
    '鄭萍':   {'dept': '沖床組', 'role': 'field_worker',       'cross_group': None},
    '蘇英輝': {'dept': '沖床組', 'role': 'field_worker',       'cross_group': None},
    '阿塘':   {'dept': '沖床組', 'role': 'field_worker',       'cross_group': None},
    '陳奕廷': {'dept': '沖床組', 'role': 'field_worker',       'cross_group': None},

    # ─── 現場：加工組（白惇維 = 代理組長，已列研發課）────────
    '張舒涵': {'dept': '加工組', 'role': 'field_leader',  'cross_group': None},
    '阮祝玲': {'dept': '加工組', 'role': 'field_leader',  'cross_group': None},
    '沙空':   {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '孫培娜': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '潘其龍': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '俗吉':   {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '宋善蓬': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '趙秋賢': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '蘇華姑': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '張政芳': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '瓦山':   {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '蘇拉碟': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '楊瑪莉': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '蔡麗花': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '戴全柏': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '班文':   {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '傅文成': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '沈怡婷': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '他那空': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '央南':   {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '匹拉':   {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '斯立查': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '林氏美': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '巴查':   {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '偉昌':   {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '努瓦':   {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},
    '潘嵐狄': {'dept': '加工組', 'role': 'field_worker',  'cross_group': None},

    # ─── 製造課直屬 ───────────────────────────────────────────
    '楊玉鳳': {'dept': '製造課', 'role': 'section_chief', 'cross_group': None},
    '鍾欣芳': {'dept': '製造課', 'role': 'admin_staff',   'cross_group': 'admin1'},
}

# ============================================================
# 評核關係計算函式
# ============================================================

def get_active_employees():
    """取得所有非排除人員"""
    return {name: info for name, info in EMPLOYEES.items() if name not in EXCLUDED}

def get_by_role(role, active=None):
    if active is None:
        active = get_active_employees()
    return [n for n, i in active.items() if i['role'] == role]

def get_by_dept(dept, active=None):
    if active is None:
        active = get_active_employees()
    return [n for n, i in active.items() if i['dept'] == dept]

def get_by_cross_group(group, active=None):
    if active is None:
        active = get_active_employees()
    return [n for n, i in active.items() if i.get('cross_group') == group]


def compute_evaluation_tasks(evaluator, active):
    """
    回傳 evaluator 需要填寫的評分表資訊：
    [{'sheet': 工作表名稱, 'evaluatees': [姓名...]}]
    """
    info = active.get(evaluator)
    if not info:
        return []

    role = info['role']
    dept = info['dept']
    cross_group = info.get('cross_group')
    tasks = []

    # ── 現場 作業員：不需填表（被評分對象）───────────────────────
    if role == 'field_worker':
        return []

    # ── 工程師：只做部內共評，不參與跨部門共評 ────────────────────
    if role == 'engineer':
        same_dept = [n for n in get_by_dept(dept, active) if n != evaluator]
        if same_dept:
            tasks.append({'sheet': '行政部內共評', 'evaluatees': same_dept})
        return tasks

    # ── 現場 班長：班長共評（同為班長互評）──────────────────────
    if role == 'field_leader':
        other_leaders = [n for n in get_by_role('field_leader', active) if n != evaluator]
        if other_leaders:
            tasks.append({'sheet': '班長共評', 'evaluatees': other_leaders})
        return tasks

    # ── 現場 組長：評自己組的作業員 + 互評其他組長 ───────────────
    if role == 'field_group_leader':
        # 1) 評自己組的班長（組長初評班長）
        own_leaders = [n for n in get_by_dept(dept, active) if active[n]['role'] == 'field_leader']
        if own_leaders:
            tasks.append({'sheet': '組長初評班長', 'evaluatees': own_leaders})

        # 2) 評自己組的作業員（現場組長初評&主管複評）
        own_workers = [n for n in get_by_dept(dept, active) if active[n]['role'] == 'field_worker']
        if own_workers:
            tasks.append({'sheet': '現場組長初評&主管複評', 'evaluatees': own_workers})

        # 3) 組長互評（其他組長）
        other_group_leaders = [n for n in get_by_role('field_group_leader', active) if n != evaluator]
        if other_group_leaders:
            tasks.append({'sheet': '組長共評', 'evaluatees': other_group_leaders})

        return tasks

    # ── 行政 一般人員：部內互評 + 跨部門共評 ─────────────────────
    if role == 'admin_staff':
        # 1) 部內互評：同部門其他人
        same_dept = [n for n in get_by_dept(dept, active) if n != evaluator]
        if same_dept:
            tasks.append({'sheet': '行政部內共評', 'evaluatees': same_dept})

        # 2) 跨部門共評：同 cross_group 的其他部門人員（只含一般行政人員與行政組長）
        if cross_group:
            cross_members = [n for n in get_by_cross_group(cross_group, active)
                             if n != evaluator
                             and active[n]['dept'] != dept
                             and active[n]['role'] in ('admin_staff', 'admin_group_leader')]
            if cross_members:
                tasks.append({'sheet': '跨部門共評', 'evaluatees': cross_members})
        return tasks

    # ── 行政 組長：評其他行政組長（行政組長互評）──────────────────
    if role == 'admin_group_leader':
        # 1) 與其他行政組長互評（使用行政部內共評表格）
        other_admin_leaders = [n for n in get_by_role('admin_group_leader', active) if n != evaluator]
        if other_admin_leaders:
            tasks.append({'sheet': '行政部內共評', 'evaluatees': other_admin_leaders})

        # 2) 跨部門共評（同 cross_group 其他部門，只含一般行政人員與行政組長）
        if cross_group:
            cross_members = [n for n in get_by_cross_group(cross_group, active)
                             if n != evaluator
                             and active[n]['dept'] != dept
                             and active[n]['role'] in ('admin_staff', 'admin_group_leader')]
            if cross_members:
                tasks.append({'sheet': '跨部門共評', 'evaluatees': cross_members})
        return tasks

    # ── 課長/副課長：課級互評 + 部內複評下屬 ─────────────────────
    if role == 'section_chief':
        # 1) 課長互評
        other_chiefs = [n for n in get_by_role('section_chief', active) if n != evaluator]
        if other_chiefs:
            tasks.append({'sheet': '課級共評', 'evaluatees': other_chiefs})

        # 2) 複評自己課的組長
        own_group_leaders = [n for n in get_by_dept(dept, active) if active[n]['role'] == 'admin_group_leader']
        if own_group_leaders:
            tasks.append({'sheet': '課級初評組長', 'evaluatees': own_group_leaders})

        # 3) 複評一般行政人員
        own_staff = [n for n in get_by_dept(dept, active) if active[n]['role'] == 'admin_staff']
        if own_staff:
            tasks.append({'sheet': '主管複評行政', 'evaluatees': own_staff})

        # 4) 若有額外負責的現場組（白惇維 = 加工組代理組長）
        extra_group = info.get('extra_field_group')
        if extra_group:
            # 初評加工組班長
            group_leaders = [n for n in get_by_dept(extra_group, active)
                             if active[n]['role'] == 'field_leader']
            if group_leaders:
                tasks.append({'sheet': '組長初評班長', 'evaluatees': group_leaders})
            # 初評加工組作業員
            group_workers = [n for n in get_by_dept(extra_group, active)
                             if active[n]['role'] == 'field_worker']
            if group_workers:
                tasks.append({'sheet': '現場組長初評&主管複評', 'evaluatees': group_workers})

        return tasks

    # ── 副理：副理/經理互評 ────────────────────────────────────────
    if role == 'deputy_manager':
        peers = [n for n in active if active[n]['role'] in ('deputy_manager', 'manager') and n != evaluator]
        if peers:
            tasks.append({'sheet': '副理 經理共評', 'evaluatees': peers})
        return tasks

    # ── 經理：副理/經理互評 ────────────────────────────────────────
    if role == 'manager':
        peers = [n for n in active if active[n]['role'] in ('deputy_manager', 'manager') and n != evaluator]
        if peers:
            tasks.append({'sheet': '副理 經理共評', 'evaluatees': peers})
        return tasks

    # ── 高層（鄭景文/陳佳倫）：由人資確認，本次跳過 ─────────────
    if role == 'top_mgmt':
        return []

    return []


# ============================================================
# Excel 工作表複製工具
# ============================================================

def copy_sheet_to_workbook(src_ws, dest_wb, sheet_title):
    """
    將來源工作表的所有內容（含格式）複製到目標活頁簿，
    回傳新建立的工作表。
    """
    dest_ws = dest_wb.create_sheet(title=sheet_title)

    # 複製欄寬、列高
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dest_ws.column_dimensions[col_letter].width = col_dim.width
    for row_num, row_dim in src_ws.row_dimensions.items():
        dest_ws.row_dimensions[row_num].height = row_dim.height

    # 複製儲存格內容與格式
    for row in src_ws.iter_rows():
        for src_cell in row:
            dest_cell = dest_ws.cell(row=src_cell.row, column=src_cell.column)
            # 值（跳過公式暫存，避免共享公式問題）
            if src_cell.value is not None:
                if isinstance(src_cell.value, str) and src_cell.value.startswith('='):
                    dest_cell.value = src_cell.value  # 保留公式
                else:
                    dest_cell.value = src_cell.value

            # 樣式
            if src_cell.has_style:
                dest_cell.font        = copy.copy(src_cell.font)
                dest_cell.border      = copy.copy(src_cell.border)
                dest_cell.fill        = copy.copy(src_cell.fill)
                dest_cell.number_format = src_cell.number_format
                dest_cell.protection  = copy.copy(src_cell.protection)
                dest_cell.alignment   = copy.copy(src_cell.alignment)

    # 複製合併儲存格
    for merged in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(merged))

    return dest_ws


def fill_evaluatee_names(ws, evaluator_name, dept, evaluatees):
    """
    在工作表中填入：
    - A5 = 部門/單位
    - B5 = 評分者姓名
    - D6, E6, F6 ... = 受評者姓名（逐一）
    """
    ws['A5'] = dept
    ws['B5'] = evaluator_name

    for idx, name in enumerate(evaluatees):
        col = 4 + idx  # D = col 4
        ws.cell(row=6, column=col, value=name)


# ============================================================
# 主程式：生成所有人的評分表
# ============================================================

def main():
    print("載入評分表範本…")
    template_wb = load_workbook(TEMPLATE_FILE, read_only=False, keep_vba=False)
    active = get_active_employees()

    skipped = []
    generated = []
    no_tasks = []

    for evaluator in sorted(active.keys()):
        info = active[evaluator]
        tasks = compute_evaluation_tasks(evaluator, active)

        if not tasks:
            no_tasks.append(f"  {evaluator}（{info['dept']} / {info['role']}）")
            continue

        # 建立新活頁簿
        new_wb = Workbook()
        new_wb.remove(new_wb.active)  # 移除預設空白頁

        for task in tasks:
            sheet_name = task['sheet']
            evaluatees = task['evaluatees']

            if sheet_name not in template_wb.sheetnames:
                print(f"  ⚠️  找不到工作表：{sheet_name}（{evaluator}）")
                continue

            src_ws = template_wb[sheet_name]
            dest_ws = copy_sheet_to_workbook(src_ws, new_wb, sheet_name)
            fill_evaluatee_names(dest_ws, evaluator, info['dept'], evaluatees)

        # 儲存
        safe_name = evaluator.replace('/', '_')
        out_path = os.path.join(OUTPUT_DIR, f"{info['dept']}_{evaluator}_評分表.xlsx")
        new_wb.save(out_path)
        generated.append(f"  ✅ {evaluator}（{info['dept']}）→ {len(tasks)} 張表 / {sum(len(t['evaluatees']) for t in tasks)} 位受評者")

    # ── 輸出摘要 ──────────────────────────────────────────────
    print(f"\n{'='*55}")
    print(f"✅ 成功生成 {len(generated)} 人的評分表")
    print(f"{'='*55}")
    for line in generated:
        print(line)

    if no_tasks:
        print(f"\n⚙️  以下人員無需填寫評分表（作業員 / 高層 / 特殊）：")
        for line in no_tasks:
            print(line)

    print(f"\n📁 輸出位置：{OUTPUT_DIR}")
    print(f"   共 {len(generated)} 個 Excel 檔案")


if __name__ == '__main__':
    main()
