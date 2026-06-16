#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
績效考核分數匯入腳本
115年度 上半年 績效考核 — 樂聯工業股份有限公司

用法：
  python3 import_scores.py <已評分的評分表資料夾> <分數總表.xlsx> <輸出檔案.xlsx>

範例：
  python3 import_scores.py ./已完成評分表/ 114_總表.xlsx 115_總表_匯入.xlsx

每年沿用方式：
  1. 將當年已填好分數的 .xlsx 評分表放入指定資料夾
  2. 準備好分數總表（可從去年複製，清空 H-AO 欄分數後使用）
  3. 執行此腳本
"""

import os
import sys
import glob
from openpyxl import load_workbook

# ============================================================
# 工作表名稱 → 分數類別
# 判斷依據：評分者身份（初評主管 → 部內；高階主管複評 → 複評；跨部門 → 跨共評）
# ============================================================
SHEET_CATEGORY = {
    '行政部內共評':             '部內',
    '跨部門共評':               '跨共評',
    '主管複評行政':             '複評',
    '副理 經理共評':            '部內',
    '組長初評班長':             '部內',
    '班長共評':                 '部內',
    '課級初評組長':             '部內',
    '組長共評':                 '部內',
    '課級複評班長':             '複評',
    '課級共評':                 '部內',
    '部級初評課級':             '部內',
    '總經理複評課級':           '複評',
    '現場組長初評&主管複評':    '部內',
}

# 各工作表的總分列（=SUM 所在列，來自範本）
SHEET_TOTAL_ROW = {
    '行政部內共評':             22,
    '跨部門共評':               18,
    '主管複評行政':             20,
    '副理 經理共評':            22,
    '組長初評班長':             18,
    '班長共評':                 18,
    '課級初評組長':             19,
    '組長共評':                 19,
    '課級複評班長':             18,
    '課級共評':                 21,
    '部級初評課級':             21,
    '總經理複評課級':           21,
    '現場組長初評&主管複評':    21,
}

# 計算成績工作表欄位範圍（1-based col index）
CATEGORY_COLS = {
    '部內':   (8,  19),   # H~S：部內分數~部內分數11（12欄）
    '跨共評': (20, 31),   # T~AE：跨共評~跨共評10（12欄）
    '複評':   (32, 41),   # AF~AO：複評~複評10（10欄）
}

TARGET_SHEET = '計算成績'
NAME_COL = 2   # B欄：姓名
DATA_START_ROW = 2
DATA_END_ROW = 146


# ============================================================
def read_scores_from_eval_file(filepath):
    """
    讀取一份已評分完畢的評分表，回傳所有分數記錄。
    回傳格式：[{'name': 受評者, 'score': float, 'category': 分數類別, 'evaluator': 評分者, 'sheet': 工作表}]
    """
    records = []
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  ⚠️  無法開啟 {os.path.basename(filepath)}: {e}")
        return records

    for sheet_name in wb.sheetnames:
        if sheet_name not in SHEET_CATEGORY:
            continue

        category  = SHEET_CATEGORY[sheet_name]
        total_row = SHEET_TOTAL_ROW[sheet_name]
        ws = wb[sheet_name]

        evaluator = ws['B5'].value or ''
        evaluator = str(evaluator).strip()

        # 從 D6 開始讀取受評者姓名與對應總分
        col = 4  # D欄
        while col <= 60:
            name_val = ws.cell(row=6, column=col).value
            if name_val is None or str(name_val).strip() == '':
                break
            name = str(name_val).strip()
            score_val = ws.cell(row=total_row, column=col).value
            if score_val is not None:
                try:
                    score = float(score_val)
                    records.append({
                        'name':      name,
                        'score':     round(score, 4),
                        'category':  category,
                        'evaluator': evaluator,
                        'sheet':     sheet_name,
                    })
                except (ValueError, TypeError):
                    print(f"  ⚠️  {sheet_name} 的 {name} 總分非數字：{score_val}")
            col += 1

    wb.close()
    return records


# ============================================================
def main(eval_folder, summary_path, output_path):
    print(f"載入分數總表：{summary_path}")
    wb_summary = load_workbook(summary_path)

    if TARGET_SHEET not in wb_summary.sheetnames:
        print(f"❌ 找不到工作表「{TARGET_SHEET}」，請確認分數總表格式。")
        sys.exit(1)

    ws_summary = wb_summary[TARGET_SHEET]

    # 建立 姓名 → 列號 的對照表
    name_to_row = {}
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        name = ws_summary.cell(row=row, column=NAME_COL).value
        if name and str(name).strip():
            name_to_row[str(name).strip()] = row

    print(f"  → 共 {len(name_to_row)} 位員工")

    # 記錄各人各類別已用欄位（避免重複填入）
    used_cols = {}
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        used_cols[row] = {'部內': set(), '跨共評': set(), '複評': set()}
        for cat, (c_start, c_end) in CATEGORY_COLS.items():
            for col in range(c_start, c_end + 1):
                val = ws_summary.cell(row=row, column=col).value
                if val is not None and str(val).strip() not in ('', '#DIV/0!', '#N/A'):
                    used_cols[row][cat].add(col)

    # 掃描評分表資料夾
    eval_files = sorted(glob.glob(os.path.join(eval_folder, '*.xlsx')))
    if not eval_files:
        print(f"⚠️  在 {eval_folder} 中找不到任何 .xlsx 檔案。")
        sys.exit(1)

    print(f"\n開始匯入（共 {len(eval_files)} 份評分表）…")

    total_filled = 0
    not_found    = []
    col_full     = []

    for filepath in eval_files:
        fname = os.path.basename(filepath)
        records = read_scores_from_eval_file(filepath)
        if not records:
            print(f"  [{fname}] 無分數資料")
            continue

        filled = 0
        for rec in records:
            name     = rec['name']
            score    = rec['score']
            category = rec['category']

            if name not in name_to_row:
                if name not in not_found:
                    not_found.append(name)
                continue

            row = name_to_row[name]
            c_start, c_end = CATEGORY_COLS[category]

            # 找下一個空欄
            next_col = None
            for col in range(c_start, c_end + 1):
                if col not in used_cols[row][category]:
                    next_col = col
                    break

            if next_col is None:
                key = f"{name}({category})"
                if key not in col_full:
                    col_full.append(key)
                continue

            ws_summary.cell(row=row, column=next_col).value = score
            used_cols[row][category].add(next_col)
            filled += 1

        total_filled += filled
        print(f"  [{fname}] 填入 {filled} 筆")

    # 儲存
    wb_summary.save(output_path)

    print(f"\n{'='*55}")
    print(f"✅ 完成！共填入 {total_filled} 筆分數")
    print(f"📁 輸出檔案：{output_path}")

    if not_found:
        print(f"\n⚠️  以下姓名在總表中找不到（{len(not_found)} 位）：")
        for n in not_found:
            print(f"   {n}")
    if col_full:
        print(f"\n⚠️  以下欄位已滿（超過上限）：")
        for k in col_full:
            print(f"   {k}")


# ============================================================
if __name__ == '__main__':
    if len(sys.argv) != 4:
        print(__doc__)
        print("用法: python3 import_scores.py <評分表資料夾> <分數總表.xlsx> <輸出檔案.xlsx>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2], sys.argv[3])
